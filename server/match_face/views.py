import face_recognition
import urllib.request as ur
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from authentication.models import User
import requests
from PIL import Image
from io import BytesIO
import numpy as np
from pathlib import Path
import openpyxl
from .models import Attendance
from django.utils import timezone
from .serializers import AttendanceDetailsSerializer

def isNullId(id):
    if not id:
        raise Exception("Please login again!")

def decodeImg(img):
    return ur.urlopen(img)

def load_img_from_cloudinary(url):
    response = requests.get(url=url)
    response.raise_for_status()
    image = Image.open(BytesIO(response.content))
    return image

def IsSameDate(last_attendance_taken):
    today = timezone.now().date()
    print(timezone.get_current_timezone())
    print(last_attendance_taken, today)
    if last_attendance_taken == today:
        raise Exception("Attendance already taken!")
    
def mark_as_present(user, latitude, longitude): 
    try:
        Attendance.objects.create(user= user, latitude= latitude, longitude= longitude)
    except Exception:
        raise Exception("Failed to mark you as attend!")

def create_sheet(month, year):
    if not Path(f"sheets/Attendance_sheet_{month}_{year}.xlsx").is_file():
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = f"Attendance_sheet_{month}_{year}"
            sheet["A1"] = "Date"
            sheet["B1"] = "Name"
            sheet["C1"] = "Email"
            sheet["D1"] = "Latitude"
            sheet["E1"] = "Longitude"
        
            workbook.save(f"sheets/Attendance_sheet_{month}_{year}.xlsx")
            return workbook, sheet
    else:
        workbook = openpyxl.load_workbook(f"sheets/Attendance_sheet_{month}_{year}.xlsx")
        if workbook:
            sheet = workbook.active
            return workbook, sheet
    
def take_attendance(data):
    curr_time = timezone.now()
    workbook, sheet = create_sheet(curr_time.month, curr_time.year)
    
    for row in data:
        sheet.append([row['time'].strftime('%d-%m-%Y'), f"{row['user__first_name']} {row['user__last_name']}", row['user__email'], row['latitude'], row['longitude']])
        
    workbook.save(f"sheets/Attendance_sheet_{curr_time.month}_{curr_time.year}.xlsx")
    
class MatchFace(APIView):
    def post(self, req):
        try:  
            user = User.objects.get(email=req.data["email"])
            IsSameDate(user.last_attendance_taken)
            decodedImg = decodeImg(req.data["img"])
            unknown_image = face_recognition.load_image_file(decodedImg)
            unknown_image_face_location = face_recognition.face_locations(unknown_image)
            
            if(unknown_image_face_location):
                userAvatar = load_img_from_cloudinary(user.img_path)
                userAvatar_np = np.array(userAvatar)
                
                known_image_face_location = face_recognition.face_locations(userAvatar_np)
                
                if known_image_face_location:
                    known_image_encoding = face_recognition.face_encodings(face_image=userAvatar_np, known_face_locations=known_image_face_location)[0]
                    unknown_image_encoding = face_recognition.face_encodings(face_image=unknown_image, known_face_locations=unknown_image_face_location)[0]
                        
                    results = face_recognition.compare_faces([known_image_encoding], unknown_image_encoding)
                    if results[0] == True:
                        user.last_attendance_taken = timezone.now()
                        user.save()
                        mark_as_present(user=user, latitude= req.data["latitude"], longitude= req.data["longitude"])
                        return Response({"res": "Face matched!"}, status=status.HTTP_200_OK)
                    
                    else:
                        return Response({"res": "Face didn't match!"}, status=status.HTTP_400_BAD_REQUEST)
                else:
                    return Response({"res": "You avatar doesn't contain any face!"}, status=status.HTTP_400_BAD_REQUEST)
            else:
                return Response({"res": "No face detected!"}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            print(e)
            return Response({"res": f"{e}"}, status=status.HTTP_400_BAD_REQUEST)
        
        
class AttendanceDetailsView(APIView):
    def get(self, req, id, format=None):
        try:
            isNullId(id=id)
            attendance_data = Attendance.objects.filter(user=id);
            if attendance_data:
                serializer = AttendanceDetailsSerializer(attendance_data, many=True)
                return Response(data=serializer.data, status=status.HTTP_200_OK)
            else:
                raise Exception("User doesn't exists!")
                
        except Exception as e:
            return Response(data={"msg": f"{e}"}, status=status.HTTP_400_BAD_REQUEST)
        
        
        
    def post(self, req, id, format=None):
        try:
            isNullId(id=id)
            all_attendance_data = Attendance.objects.select_related('user').values( 'time','user__first_name', 'user__last_name', 'user__email', 'latitude', 'longitude')
            attendance_list = list(all_attendance_data)
            take_attendance(data=attendance_list)
            Attendance.objects.all().delete()
            return Response(data={"res": "Data added to execl file, check the sheets directory in your project!"}, status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response(data={"res": f"{e}"}, status=status.HTTP_400_BAD_REQUEST)